#!/usr/bin/env python3
"""
PLEX² – moteur DOE Explorer
Basé sur le script source utilisateur DoE_Explorer_PhAm-ForApp.py.
Version Windows / GUI avec LHS automatique et résumé Excel ordonné.
"""

from __future__ import annotations

import argparse
import itertools
import math
import re
import sys
from collections import Counter, defaultdict
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import pyDOE3
except ImportError as exc:
    raise ImportError("pyDOE3 n'est pas installe. Installer avec : pip install pyDOE3") from exc


# ============================================================
# COULEURS PAR CATEGORIE
# ============================================================
CATEGORY_COLORS: dict[str, tuple[str, str]] = {
    "resume":       ("2E2E2E", "FFFFFF"),
    "complet":      ("1F4E79", "FFFFFF"),
    "fraction":     ("833C00", "FFFFFF"),
    "plackett":     ("7030A0", "FFFFFF"),
    "frac2L":       ("BF5F00", "FFFFFF"),
    "dsd":          ("0A6B6B", "FFFFFF"),
    "ofat":         ("3A3A3A", "FFFFFF"),
    "taguchi":      ("375623", "FFFFFF"),
    "gsd":          ("1D6B55", "FFFFFF"),
    "ccd":          ("006B6B", "FFFFFF"),
    "ccd_coded":    ("4BACC6", "FFFFFF"),
    "box-behnken":  ("C00000", "FFFFFF"),
    "bb_coded":     ("FF7B7B", "000000"),
    "doehlert":     ("4A235A", "FFFFFF"),
    "doehlert_cod": ("9B59B6", "FFFFFF"),
    "lhs":          ("7F6000", "FFFFFF"),
    "spacefill":    ("0F766E", "FFFFFF"),
}


def _category(type_label: str) -> str:
    t = type_label.lower()
    if "resume" in t or "summary" in t:
        return "resume"
    if "min/max" in t or "reformat" in t:
        return "frac2L"
    if "dsd" in t or "definitive screening" in t:
        return "dsd"
    if "ofat" in t or "one-factor" in t:
        return "ofat"
    if "complet" in t:
        return "complet"
    if "fraction" in t or "frac" in t:
        return "fraction"
    if "plackett" in t:
        return "plackett"
    if "taguchi" in t:
        return "taguchi"
    if "gsd" in t:
        return "gsd"
    if "ccd" in t or "composite" in t:
        return "ccd_coded" if "coded" in t else "ccd"
    if "box" in t:
        return "bb_coded" if "coded" in t else "box-behnken"
    if "doehlert" in t:
        return "doehlert_cod" if "coded" in t else "doehlert"
    if "space" in t or "maximin" in t or "d-optimal" in t or "d optimal" in t:
        return "spacefill"
    if "lhs" in t or "hypercube" in t:
        return "lhs"
    return "resume"


# ============================================================
# HELPERS NOMS DE FEUILLE
# ============================================================
_INVALID = re.compile(r"[\\/:*?\[\]]")


def _safe_name(name: str, used: set[str]) -> str:
    name = _INVALID.sub("_", name).strip()[:31] or "Sheet"
    base, i = name, 2
    while name in used:
        suffix = f"_{i}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(name)
    return name


# ============================================================
# CONVERSIONS CODE -> VALEURS REELLES
# ============================================================

def _df_full_factorial(factors: Dict[str, List[Any]]) -> pd.DataFrame:
    names = list(factors.keys())
    grid = list(itertools.product(*(factors[n] for n in names)))
    df = pd.DataFrame(grid, columns=names)
    df.insert(0, "Run", range(1, len(df) + 1))
    return df



def _df_from_0indexed(mat: np.ndarray, factors: Dict[str, List[Any]]) -> pd.DataFrame:
    names = list(factors.keys())
    data = {n: [factors[n][int(c)] for c in mat[:, j]] for j, n in enumerate(names)}
    df = pd.DataFrame(data)
    df.insert(0, "Run", range(1, len(df) + 1))
    return df



def _df_from_pm1_2level(mat: np.ndarray, factors: Dict[str, List[Any]]) -> pd.DataFrame:
    names = list(factors.keys())
    data = {
        n: [factors[n][0] if v < 0 else factors[n][1] for v in mat[:, j]]
        for j, n in enumerate(names)
    }
    df = pd.DataFrame(data)
    df.insert(0, "Run", range(1, len(df) + 1))
    return df



def _df_from_rsm(mat: np.ndarray, factors: Dict[str, List[Any]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    names = list(factors.keys())
    df_coded = pd.DataFrame(mat, columns=[f"{n}_coded" for n in names])
    df_coded.insert(0, "Run", range(1, len(df_coded) + 1))

    def _is_num(lvls: list[Any]) -> bool:
        return all(isinstance(x, (int, float, np.number)) for x in lvls)

    if not all(_is_num(factors[n]) for n in names):
        return df_coded, df_coded

    mins = np.array([min(factors[n]) for n in names], dtype=float)
    maxs = np.array([max(factors[n]) for n in names], dtype=float)
    centers = (mins + maxs) / 2.0
    scales = (maxs - mins) / 2.0
    scales[scales == 0] = 1.0

    real = centers + mat * scales
    df_real = pd.DataFrame(real, columns=names)
    df_real.insert(0, "Run", range(1, len(df_real) + 1))
    return df_real, df_coded


# ============================================================
# HELPERS DSD / LHS / TRI
# ============================================================

def _is_numeric_factor(lvls: List[Any]) -> bool:
    return all(isinstance(v, (int, float, np.number)) for v in lvls)



def _hadamard_sylvester(n: int) -> np.ndarray:
    H = np.array([[1.0]])
    while H.shape[0] < n:
        H = np.block([[H, H], [H, -H]])
    return H



def _dsd_matrix(k: int) -> np.ndarray:
    n = 1
    while n <= k:
        n *= 2
    H = _hadamard_sylvester(n)
    sub = H[1 : k + 1, 1 : k + 1].copy()
    np.fill_diagonal(sub, 0.0)
    return np.vstack([sub, -sub, np.zeros((1, k))])



def _compute_auto_lhs_samples(factors: Dict[str, List[Any]]) -> int:
    """Calcule une taille LHS bornée pour éviter les volumes excessifs."""
    numeric_profile = [len(v) for v in factors.values() if _is_numeric_factor(v)]
    if not numeric_profile:
        return 0

    k_num = len(numeric_profile)
    avg_l = sum(numeric_profile) / k_num
    max_l = max(numeric_profile)

    full_runs = 1
    for n in numeric_profile:
        full_runs *= n
        if full_runs >= 100_000:
            full_runs = 100_000
            break

    base = max(12, 4 * k_num)
    coverage = math.ceil(1.5 * k_num * avg_l)
    spread = math.ceil(2.25 * max_l * math.sqrt(k_num))
    root_full = math.ceil(math.sqrt(full_runs))
    ceiling = max(24, min(300, 20 * k_num))

    candidate = max(base, coverage, spread, root_full)
    candidate = min(candidate, ceiling)
    if candidate % 2:
        candidate += 1
    return max(12, candidate)






# ============================================================
# ============================================================
# SPACE-FILLING MAXIMIN DISCRET - niveaux utilisateur exacts
# ============================================================

def _spacefill_domain_size(factors: Dict[str, List[Any]]) -> int:
    """Nombre de configurations uniques possibles avec les niveaux utilisateur."""
    total = 1
    for lvls in factors.values():
        total *= len(lvls)
    return total


def _spacefill_candidates(
    factors: Dict[str, List[Any]],
    n_samples: int,
    rng: np.random.Generator,
    max_candidates: int = 200_000,
) -> Tuple[pd.DataFrame, int, bool]:
    """Pool candidat discret : produit complet si raisonnable, sinon pool unique tiré sur les niveaux."""
    names = list(factors.keys())
    total = _spacefill_domain_size(factors)

    if total <= max_candidates:
        grid = itertools.product(*(factors[n] for n in names))
        return pd.DataFrame(grid, columns=names), total, True

    # Domaine énorme : pool aléatoire unique, sans aucune valeur intermédiaire.
    target = min(total, max(n_samples * 12, 5_000))
    seen: set[tuple[Any, ...]] = set()
    rows: list[tuple[Any, ...]] = []
    tries = 0
    while len(rows) < target and tries < target * 25:
        row = tuple(factors[n][int(rng.integers(0, len(factors[n])))] for n in names)
        if row not in seen:
            seen.add(row)
            rows.append(row)
        tries += 1

    if len(rows) < n_samples:
        raise ValueError(f"pool discret insuffisant : {len(rows)} configurations uniques pour {n_samples} demandées")
    return pd.DataFrame(rows, columns=names), total, False


def _spacefill_encode(df: pd.DataFrame, factors: Dict[str, List[Any]]) -> np.ndarray:
    """Encodage distance : numérique normalisé, catégoriel one-hot."""
    cols: list[np.ndarray] = []
    for name, lvls in factors.items():
        if _is_numeric_factor(lvls):
            lo, hi = float(min(lvls)), float(max(lvls))
            vals = df[name].astype(float).to_numpy()
            x = np.zeros(len(df), dtype=np.float32) if math.isclose(lo, hi) else ((vals - lo) / (hi - lo)).astype(np.float32)
            cols.append(x.reshape(-1, 1))
        else:
            values = df[name].astype(str).to_numpy()
            levels = [str(v) for v in lvls]
            mapping = {v: i for i, v in enumerate(levels)}
            onehot = np.zeros((len(df), len(levels)), dtype=np.float32)
            for r, value in enumerate(values):
                onehot[r, mapping.get(str(value), 0)] = 1.0
            if len(levels) > 1:
                onehot /= math.sqrt(2.0)
            cols.append(onehot)
    return np.column_stack(cols).astype(np.float32) if cols else np.empty((len(df), 0), dtype=np.float32)


def _spacefill_greedy_maximin(E: np.ndarray, n_samples: int) -> np.ndarray:
    """
    Farthest-point sampling maximin très rapide.
    Complexité ≈ O(n_samples × n_candidates × dimensions), sans échange local coûteux.
    """
    m = E.shape[0]
    selected = np.empty(n_samples, dtype=np.int64)
    selected_mask = np.zeros(m, dtype=bool)

    center = np.mean(E, axis=0)
    current = int(np.argmin(np.einsum("ij,ij->i", E - center, E - center)))
    min_d2 = np.full(m, np.inf, dtype=np.float32)

    for t in range(n_samples):
        selected[t] = current
        selected_mask[current] = True
        diff = E - E[current]
        d2 = np.einsum("ij,ij->i", diff, diff).astype(np.float32)
        min_d2 = np.minimum(min_d2, d2)
        min_d2[selected_mask] = -1.0
        if t < n_samples - 1:
            current = int(np.argmax(min_d2))
    return selected


def _spacefill_min_distance(E: np.ndarray, selected_idx: np.ndarray) -> float:
    """Distance minimale du plan sélectionné, pour la note résumé."""
    if len(selected_idx) <= 1:
        return 0.0
    sample = E[selected_idx]
    best = np.inf
    block = 512
    for i in range(0, len(sample), block):
        part = sample[i:i + block]
        d2 = ((part[:, None, :] - sample[None, :, :]) ** 2).sum(axis=2)
        for r in range(len(part)):
            d2[r, i + r] = np.inf
        best = min(best, float(np.min(d2)))
    return math.sqrt(best) if np.isfinite(best) else 0.0


def _spacefill_projection_encode(df: pd.DataFrame, factors: Dict[str, List[Any]]) -> np.ndarray:
    """Encodage compact par facteur pour les critères de projection."""
    cols: list[np.ndarray] = []
    for name, lvls in factors.items():
        if _is_numeric_factor(lvls):
            lo, hi = float(min(lvls)), float(max(lvls))
            vals = df[name].astype(float).to_numpy()
            x = np.zeros(len(df), dtype=np.float64) if math.isclose(lo, hi) else ((vals - lo) / (hi - lo)).astype(np.float64)
        else:
            mapping = {_value_key(value): idx for idx, value in enumerate(lvls)}
            denom = max(1, len(lvls) - 1)
            x = np.array([mapping.get(_value_key(value), 0) / denom for value in df[name].tolist()], dtype=np.float64)
        cols.append(x.reshape(-1, 1))
    return np.column_stack(cols) if cols else np.empty((len(df), 0), dtype=np.float64)


def _spacefill_greedy_maxpro(P: np.ndarray, n_samples: int) -> np.ndarray:
    """
    Sélection gloutonne inspirée MaxPro.
    On minimise la somme des inverses des distances projetées déjà créées.
    """
    m = P.shape[0]
    if P.shape[1] == 0:
        return np.arange(min(n_samples, m), dtype=np.int64)

    selected = np.empty(n_samples, dtype=np.int64)
    selected_mask = np.zeros(m, dtype=bool)
    center = np.full(P.shape[1], 0.5, dtype=np.float64)
    current = int(np.argmin(np.einsum("ij,ij->i", P - center, P - center)))
    log_penalty = np.full(m, -np.inf, dtype=np.float64)
    eps = 1e-3

    for t in range(n_samples):
        selected[t] = current
        selected_mask[current] = True
        diff = np.maximum(np.abs(P - P[current]), eps)
        pair_log_penalty = -2.0 * np.sum(np.log(diff), axis=1)
        log_penalty = np.logaddexp(log_penalty, pair_log_penalty)
        log_penalty[selected_mask] = np.inf
        if t < n_samples - 1:
            current = int(np.argmin(log_penalty))
    return selected


def _space_filling_maximin_design(
    factors: Dict[str, List[Any]],
    n_samples: int,
    random_state: int = 20260519,
) -> Tuple[pd.DataFrame, str]:
    """Plan Space-Filling discret maximin, sans valeurs intermédiaires."""
    if n_samples < 1:
        raise ValueError("Le nombre d'expériences du plan space-filling doit être positif.")

    total = _spacefill_domain_size(factors)
    if n_samples > total:
        raise ValueError(f"{n_samples} essais demandés > {total} configurations uniques possibles avec les niveaux utilisateur.")

    rng = np.random.default_rng(random_state)
    candidates, total, exhaustive = _spacefill_candidates(factors, n_samples, rng)

    if n_samples == total and exhaustive:
        df = candidates.reset_index(drop=True).copy()
        df.insert(0, "Run", range(1, len(df) + 1))
        return df, f"Plan discret complet : toutes les {total} configurations uniques sont utilisées. Niveaux exacts utilisateur."

    E = _spacefill_encode(candidates, factors)
    selected_idx = _spacefill_greedy_maximin(E, n_samples)
    min_dist = _spacefill_min_distance(E, selected_idx)

    df = candidates.iloc[selected_idx].reset_index(drop=True)
    df.insert(0, "Run", range(1, len(df) + 1))
    note = (
        f"Plan discret sur les niveaux exacts utilisateur. Taille demandée = {n_samples}. "
        f"Domaine = {total} configurations uniques. "
        f"Pool candidat = {len(candidates)} ({'exhaustif' if exhaustive else 'échantillon discret unique'}). "
        f"Critère space-filling maximin rapide. Distance minimale = {min_dist:.4g}."
    )
    return df, note


def _space_filling_maxpro_design(
    factors: Dict[str, List[Any]],
    n_samples: int,
    random_state: int = 20260523,
) -> Tuple[pd.DataFrame, str]:
    """Plan Space-Filling discret orienté sous-projections, sans valeurs intermédiaires."""
    if n_samples < 1:
        raise ValueError("Le nombre d'expériences du plan space-filling doit être positif.")

    total = _spacefill_domain_size(factors)
    if n_samples > total:
        raise ValueError(f"{n_samples} essais demandés > {total} configurations uniques possibles avec les niveaux utilisateur.")

    rng = np.random.default_rng(random_state)
    candidates, total, exhaustive = _spacefill_candidates(factors, n_samples, rng)

    if n_samples == total and exhaustive:
        df = candidates.reset_index(drop=True).copy()
        df.insert(0, "Run", range(1, len(df) + 1))
        return df, f"Plan discret complet : toutes les {total} configurations uniques sont utilisées. Niveaux exacts utilisateur."

    P = _spacefill_projection_encode(candidates, factors)
    selected_idx = _spacefill_greedy_maxpro(P, n_samples)
    E = _spacefill_encode(candidates, factors)
    min_dist = _spacefill_min_distance(E, selected_idx)

    df = candidates.iloc[selected_idx].reset_index(drop=True)
    df.insert(0, "Run", range(1, len(df) + 1))
    note = (
        f"Plan discret sur les niveaux exacts utilisateur. Taille demandée = {n_samples}. "
        f"Domaine = {total} configurations uniques. "
        f"Pool candidat = {len(candidates)} ({'exhaustif' if exhaustive else 'échantillon discret unique'}). "
        f"Critère space-filling projection maximin (MaxPro discret), favorise les sous-projections. "
        f"Distance minimale = {min_dist:.4g}."
    )
    return df, note


def _normalize_warning_note(type_label: str, notes: str, is_generated: bool) -> str:
    text = (notes or '').strip()
    lower = f"{type_label} {text}".lower()
    if not is_generated:
        if not text.lower().startswith("⚠ omis"):
            text = f"⚠ OMIS – {text.removeprefix('OMIS – ').strip()}" if text else "⚠ OMIS"
        return text
    if "latin hypercube" in lower and not text.lower().startswith("⚠ lhs"):
        text = f"⚠ LHS : {text}" if text else "⚠ LHS"
    elif ("reformat" in lower or "min/max" in lower or "exclus" in lower) and not text.startswith("⚠"):
        text = f"⚠ REFORMATAGE : {text}" if text else "⚠ REFORMATAGE"
    return text



def _classify_priority(type_label: str, notes: str, is_generated: bool) -> int:
    lower = f"{type_label} {notes}".lower()
    if not is_generated or lower.startswith("⚠ omis") or "⚠ omis" in lower:
        return 2
    if "⚠" in notes or "reformat" in lower or "min/max" in lower or " latin hypercube" in lower:
        return 1
    return 0



def _validate_plan_df(df: pd.DataFrame, type_label: str) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame):
        raise TypeError(f"Le plan '{type_label}' n'a pas produit de DataFrame pandas valide.")
    if df.empty:
        raise ValueError(f"Le plan '{type_label}' a produit un DataFrame vide.")
    if df.columns.duplicated().any():
        raise ValueError(f"Le plan '{type_label}' contient des colonnes dupliquées.")
    return df.copy()



def _build_short_label(seed: str) -> str:
    seed = seed.upper()
    seed = seed.replace(" ", "_")
    seed = seed.replace("-", "_")
    seed = re.sub(r"[^A-Z0-9_]+", "", seed)
    seed = re.sub(r"_+", "_", seed).strip("_")
    return seed[:24] or "PLAN"


# ============================================================
# TAGUCHI : AUTO-DECOUVERTE COMPLETE
# ============================================================

def _taguchi_all_compatible(factors: Dict[str, List[Any]]) -> List[Tuple[str, pd.DataFrame]]:
    try:
        from pyDOE3 import get_orthogonal_array, list_orthogonal_arrays
    except ImportError:
        return []

    needed = Counter(len(lvls) for lvls in factors.values())
    factors_by_level: dict[int, list[str]] = defaultdict(list)
    for name, lvls in factors.items():
        factors_by_level[len(lvls)].append(name)

    results: List[Tuple[str, pd.DataFrame]] = []
    for oa_name in list_orthogonal_arrays():
        try:
            oa = get_orthogonal_array(oa_name)
        except Exception:
            continue

        cols_by_level: dict[int, list[int]] = defaultdict(list)
        for j in range(oa.shape[1]):
            cols_by_level[len(np.unique(oa[:, j]))].append(j)

        if not all(len(cols_by_level.get(lvl, [])) >= req for lvl, req in needed.items()):
            continue

        data: dict[str, list[Any]] = {}
        for lvl, fnames in factors_by_level.items():
            for fname, col_idx in zip(fnames, cols_by_level[lvl]):
                lvls = factors[fname]
                data[fname] = [lvls[int(c)] for c in oa[:, col_idx].astype(int)]

        df = pd.DataFrame(data)
        df.insert(0, "Run", range(1, len(df) + 1))
        results.append((oa_name, df))

    results.sort(key=lambda x: (len(x[1]), x[0]))
    return results


# ============================================================
# FORMATAGE EXCEL
# ============================================================

SPATIAL_CELL_PIXELS = 34
FRACTAL_DECAY = 0.62
SPATIAL_MAX_LEVELS_PER_FACTOR = 6
SPATIAL_MAX_GRID_CELLS = 20_000
SPATIAL_AXIS_MIN_COLUMN_WIDTH = 9
SPATIAL_AXIS_MAX_COLUMN_WIDTH = 30
SPATIAL_AXIS_MIN_ROW_HEIGHT = 30
SPATIAL_AXIS_MAX_ROW_HEIGHT = 90


def _is_run_column(column: Any) -> bool:
    return str(column).strip().casefold() == "run"


def _is_full_factorial_plan(type_label: str) -> bool:
    return type_label.strip().casefold() == "factoriel complet"


def _value_key(value: Any) -> tuple[str, str]:
    return type(value).__name__, repr(value)


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _format_visual_value(value: Any, max_len: int = 36) -> str:
    if _is_blank_value(value):
        text = "(vide)"
    elif isinstance(value, (float, np.floating)):
        text = str(int(value)) if math.isclose(float(value), round(float(value))) else f"{float(value):.4g}"
    else:
        text = str(value)
    return text if len(text) <= max_len else f"{text[: max_len - 3]}..."


def _unique_nonblank_values(series: pd.Series) -> list[Any]:
    seen: set[tuple[str, str]] = set()
    values: list[Any] = []
    for value in series.tolist():
        if _is_blank_value(value):
            continue
        key = _value_key(value)
        if key in seen:
            continue
        seen.add(key)
        values.append(value)
    return values or ["(vide)"]


def _all_numeric(values: list[Any]) -> bool:
    return all(isinstance(value, (int, float, np.number)) and not isinstance(value, (bool, np.bool_)) for value in values)


def _build_numeric_axis(column: Any, values: list[Any], max_levels: int) -> dict[str, Any]:
    numeric_values = sorted(float(value) for value in values)
    compact_values: list[float] = []
    for value in numeric_values:
        if not compact_values or not math.isclose(value, compact_values[-1]):
            compact_values.append(value)

    if len(compact_values) <= max_levels:
        return {
            "column": column,
            "name": str(column),
            "kind": "numeric",
            "levels": compact_values,
            "labels": [_format_visual_value(value) for value in compact_values],
        }

    bins = np.array_split(np.array(compact_values, dtype=float), max_levels)
    ranges: list[tuple[float, float]] = []
    labels: list[str] = []
    for part in bins:
        lo = float(part[0])
        hi = float(part[-1])
        ranges.append((lo, hi))
        labels.append(_format_visual_value(lo) if math.isclose(lo, hi) else f"{_format_visual_value(lo)} - {_format_visual_value(hi)}")
    return {
        "column": column,
        "name": str(column),
        "kind": "numeric_bins",
        "levels": ranges,
        "labels": labels,
    }


def _build_categorical_axis(column: Any, values: list[Any], max_levels: int) -> dict[str, Any]:
    if len(values) <= max_levels:
        return {
            "column": column,
            "name": str(column),
            "kind": "categorical",
            "levels": values,
            "labels": [_format_visual_value(value) for value in values],
            "index": {_value_key(value): idx for idx, value in enumerate(values)},
        }

    chunk_size = math.ceil(len(values) / max_levels)
    groups = [values[idx: idx + chunk_size] for idx in range(0, len(values), chunk_size)]
    index: dict[tuple[str, str], int] = {}
    labels: list[str] = []
    for group_idx, group in enumerate(groups):
        for value in group:
            index[_value_key(value)] = group_idx
        labels.append(_format_visual_value(group[0]) if len(group) == 1 else f"{_format_visual_value(group[0])}...{_format_visual_value(group[-1])}")
    return {
        "column": column,
        "name": str(column),
        "kind": "categorical_groups",
        "levels": groups,
        "labels": labels,
        "index": index,
    }


def _build_visual_axis(column: Any, series: pd.Series, max_levels: int) -> dict[str, Any]:
    values = _unique_nonblank_values(series)
    if _all_numeric(values):
        return _build_numeric_axis(column, values, max_levels)
    return _build_categorical_axis(column, values, max_levels)


def _axis_level_count(axis: dict[str, Any]) -> int:
    return max(1, len(axis["labels"]))


def _axis_value_index(axis: dict[str, Any], value: Any) -> int:
    if _is_blank_value(value):
        return 0
    if axis["kind"] == "numeric":
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return 0
        for idx, level in enumerate(axis["levels"]):
            if math.isclose(numeric_value, float(level), rel_tol=1e-9, abs_tol=1e-9):
                return idx
        return int(np.argmin([abs(numeric_value - float(level)) for level in axis["levels"]]))
    if axis["kind"] == "numeric_bins":
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return 0
        for idx, (_lo, hi) in enumerate(axis["levels"]):
            if numeric_value <= hi or idx == len(axis["levels"]) - 1:
                return idx
        return len(axis["levels"]) - 1
    return int(axis.get("index", {}).get(_value_key(value), 0))


def _axis_strides(axes: list[dict[str, Any]]) -> tuple[list[int], int]:
    strides: list[int] = []
    size = 1
    for axis in axes:
        strides.append(size)
        size *= _axis_level_count(axis)
    return strides, max(1, size)


def _axis_grid_cell_count(axes: list[dict[str, Any]]) -> int:
    total = 1
    for axis in axes:
        total *= _axis_level_count(axis)
    return total


def _build_visual_axes(df: pd.DataFrame) -> list[dict[str, Any]]:
    factor_columns = [col for col in df.columns if not _is_run_column(col)]
    if not factor_columns:
        return []

    max_levels = SPATIAL_MAX_LEVELS_PER_FACTOR
    axes = [_build_visual_axis(col, df[col], max_levels) for col in factor_columns]
    if _axis_grid_cell_count(axes) <= SPATIAL_MAX_GRID_CELLS:
        return axes

    max_levels = max(2, int(SPATIAL_MAX_GRID_CELLS ** (1 / len(factor_columns))))
    while max_levels > 2:
        axes = [_build_visual_axis(col, df[col], max_levels) for col in factor_columns]
        if _axis_grid_cell_count(axes) <= SPATIAL_MAX_GRID_CELLS:
            return axes
        max_levels -= 1
    return [_build_visual_axis(col, df[col], 2) for col in factor_columns]


def _encode_visualization_column(series: pd.Series) -> np.ndarray:
    numeric = pd.to_numeric(series, errors="coerce")
    has_bool = any(isinstance(v, (bool, np.bool_)) for v in series.dropna().tolist())

    if not has_bool and numeric.notna().all():
        values = numeric.astype(float).to_numpy()
        lo = float(np.min(values))
        hi = float(np.max(values))
        if math.isclose(lo, hi):
            return np.zeros(len(series), dtype=float)
        return ((values - lo) / (hi - lo) * 2.0) - 1.0

    mapping: dict[tuple[str, str], int] = {}
    encoded: list[int] = []
    for value in series.tolist():
        key = _value_key(value)
        if key not in mapping:
            mapping[key] = len(mapping)
        encoded.append(mapping[key])

    if len(mapping) <= 1:
        return np.zeros(len(series), dtype=float)
    values = np.array(encoded, dtype=float)
    return (values / (len(mapping) - 1) * 2.0) - 1.0


def _encode_plan_for_visualization(df: pd.DataFrame) -> np.ndarray:
    factor_columns = [col for col in df.columns if not _is_run_column(col)]
    if not factor_columns:
        return np.zeros((len(df), 0), dtype=float)
    return np.column_stack([_encode_visualization_column(df[col]) for col in factor_columns])


def _scale_projection_axis(values: np.ndarray, side: int) -> np.ndarray:
    lo = float(np.min(values))
    hi = float(np.max(values))
    if math.isclose(lo, hi):
        return np.full(len(values), side // 2, dtype=int)
    scaled = np.rint((values - lo) / (hi - lo) * (side - 1)).astype(int)
    return np.clip(scaled, 0, side - 1)


def _nearest_free_cell(row: int, col: int, occupied: set[tuple[int, int]], side: int) -> tuple[int, int]:
    if (row, col) not in occupied:
        return row, col

    best: tuple[int, int, int, int] | None = None
    for radius in range(1, side + 1):
        for dr in range(-radius, radius + 1):
            for dc in range(-radius, radius + 1):
                if max(abs(dr), abs(dc)) != radius:
                    continue
                nr = row + dr
                nc = col + dc
                if not (0 <= nr < side and 0 <= nc < side) or (nr, nc) in occupied:
                    continue
                score = (dr * dr + dc * dc, abs(dr) + abs(dc), nr, nc)
                if best is None or score < best:
                    best = score
        if best is not None:
            return best[2], best[3]
    raise ValueError("Grille de visualisation saturée.")


def _run_label(value: Any, fallback: int) -> str:
    if value is None:
        return f"E{fallback}"
    try:
        number = float(value)
        if math.isnan(number):
            return f"E{fallback}"
        if math.isclose(number, round(number)):
            return f"E{int(round(number))}"
    except (TypeError, ValueError):
        pass
    return f"E{value}"


def project_fractal_2d(df: pd.DataFrame) -> dict[str, Any]:
    """
    Projection fractale lisible N dimensions -> grille 2D.
    Les dimensions 1, 3, 5... structurent l'axe horizontal par échelles imbriquées.
    Les dimensions 2, 4, 6... structurent l'axe vertical par échelles imbriquées.
    """
    n_runs = len(df)
    if n_runs <= 0:
        return {"placements": [], "x_axes": [], "y_axes": [], "x_strides": [], "y_strides": [], "width": 0, "height": 0}

    axes = _build_visual_axes(df)
    x_axes = axes[0::2]
    y_axes = axes[1::2]
    x_strides, width = _axis_strides(x_axes)
    y_strides, height = _axis_strides(y_axes)
    run_column = next((col for col in df.columns if _is_run_column(col)), None)
    run_values = df[run_column].tolist() if run_column is not None else list(range(1, n_runs + 1))

    placements: list[tuple[int, int, str, int]] = []
    for index, (_, run_value) in enumerate(zip(df.itertuples(index=False), run_values), start=1):
        row_values = df.iloc[index - 1]
        x = sum(_axis_value_index(axis, row_values[axis["column"]]) * stride for axis, stride in zip(x_axes, x_strides))
        y = sum(_axis_value_index(axis, row_values[axis["column"]]) * stride for axis, stride in zip(y_axes, y_strides))
        placements.append((height - 1 - y, x, _run_label(run_value, index), index))

    return {
        "placements": placements,
        "x_axes": x_axes,
        "y_axes": y_axes,
        "x_strides": x_strides,
        "y_strides": y_strides,
        "width": width,
        "height": height,
    }


def _contrast_font_color(hex_color: str) -> str:
    red = int(hex_color[0:2], 16)
    green = int(hex_color[2:4], 16)
    blue = int(hex_color[4:6], 16)
    luminance = (0.299 * red) + (0.587 * green) + (0.114 * blue)
    return "000000" if luminance > 150 else "FFFFFF"


def _blend_hex_colors(start_color: str, end_color: str, amount: float) -> str:
    amount = max(0.0, min(1.0, amount))
    start_rgb = [int(start_color[i:i + 2], 16) for i in (0, 2, 4)]
    end_rgb = [int(end_color[i:i + 2], 16) for i in (0, 2, 4)]
    blended = [round(start + (end - start) * amount) for start, end in zip(start_rgb, end_rgb)]
    return "".join(f"{value:02X}" for value in blended)


def _trial_sequence_color(base_color: str, order: int, n_runs: int) -> str:
    if n_runs <= 1:
        ratio = 1.0
    else:
        ratio = (order - 1) / (n_runs - 1)
    return _blend_hex_colors("FFFFFF", base_color, 0.22 + (0.78 * ratio))


def _axis_text_width(text_len: int) -> float:
    return min(
        SPATIAL_AXIS_MAX_COLUMN_WIDTH,
        max(SPATIAL_AXIS_MIN_COLUMN_WIDTH, text_len * 0.9 + 2),
    )


def _axis_text_height(text_len: int) -> float:
    return min(
        SPATIAL_AXIS_MAX_ROW_HEIGHT,
        max(SPATIAL_AXIS_MIN_ROW_HEIGHT, text_len * 3.2 + 10),
    )


def _add_spatial_visualization(ws, type_label: str, max_runs: int, base_color: str) -> None:
    if _is_full_factorial_plan(type_label):
        return

    table_max_row = ws.max_row
    table_max_column = ws.max_column
    n_runs = max(0, table_max_row - 1)
    if n_runs == 0 or n_runs > max_runs:
        return

    columns = [ws.cell(row=1, column=col).value for col in range(1, table_max_column + 1)]
    rows = list(
        ws.iter_rows(
            min_row=2,
            max_row=table_max_row,
            min_col=1,
            max_col=table_max_column,
            values_only=True,
        )
    )
    df = pd.DataFrame(rows, columns=columns)
    layout = project_fractal_2d(df)
    placements = layout["placements"]
    if not placements:
        return

    start_row = 2
    start_col = table_max_column + 2
    x_axes = layout["x_axes"]
    y_axes = layout["y_axes"]
    x_strides = layout["x_strides"]
    y_strides = layout["y_strides"]
    width = layout["width"]
    height = layout["height"]
    left_cols = max(1, len(y_axes))
    title_row = start_row
    grid_top = start_row + 1
    data_left = start_col + left_cols
    grid_bottom = grid_top + height - 1
    x_label_start = grid_bottom + 1
    end_row = x_label_start + max(1, len(x_axes)) - 1
    end_col = data_left + width - 1
    cell_height = SPATIAL_CELL_PIXELS * 0.75
    cell_width = max(2.5, (SPATIAL_CELL_PIXELS - 5) / 7)
    center = Alignment(horizontal="center", vertical="center")
    label_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    y_label_alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    x_level_alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
    label_font = Font(name="Arial", bold=True, color="1F1F1F", size=10)
    x_level_font = Font(name="Arial", bold=True, color="1F1F1F", size=9)
    title_font = Font(name="Arial", bold=True, color="1F1F1F", size=11)
    grid_border = Border(
        left=Side(style="thin", color="1F1F1F"),
        right=Side(style="thin", color="1F1F1F"),
        top=Side(style="thin", color="1F1F1F"),
        bottom=Side(style="thin", color="1F1F1F"),
    )
    trial_border = Border(
        left=Side(style="thin", color="1F1F1F"),
        right=Side(style="thin", color="1F1F1F"),
        top=Side(style="thin", color="1F1F1F"),
        bottom=Side(style="thin", color="1F1F1F"),
    )

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[title_row].height = 24
    for row in range(grid_top, grid_bottom + 1):
        ws.row_dimensions[row].height = cell_height

    for axis_idx, axis in enumerate(x_axes):
        row = x_label_start + axis_idx
        max_label_len = max([len(axis["name"])] + [len(str(label)) for label in axis["labels"]])
        ws.row_dimensions[row].height = _axis_text_height(max_label_len)

    for axis_idx, axis in enumerate(y_axes):
        col = start_col + (len(y_axes) - 1 - axis_idx)
        max_label_len = max(len(f"{axis['name']}: {label}") for label in axis["labels"])
        ws.column_dimensions[get_column_letter(col)].width = _axis_text_width(max_label_len)

    if not y_axes:
        max_name_len = max((len(axis["name"]) for axis in x_axes), default=10)
        ws.column_dimensions[get_column_letter(start_col)].width = _axis_text_width(max_name_len)

    for col in range(data_left, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = cell_width

    if end_col > start_col:
        ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
    title_cell = ws.cell(row=title_row, column=start_col)
    title_cell.value = f"Visualisation - {type_label}"
    title_cell.font = title_font
    title_cell.alignment = label_alignment
    title_cell.border = grid_border

    for row in range(grid_top, grid_bottom + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = label_alignment
            cell.border = grid_border

    for row in range(x_label_start, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = label_alignment
            cell.border = grid_border

    def _merge_if_needed(row_start: int, col_start: int, row_end: int, col_end: int) -> None:
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end)

    for axis_idx, axis in enumerate(y_axes):
        col = start_col + (len(y_axes) - 1 - axis_idx)
        stride = y_strides[axis_idx]
        level_count = _axis_level_count(axis)
        period = stride * level_count
        for group_start in range(0, height, period):
            for level_idx, level_label in enumerate(axis["labels"]):
                logical_start = group_start + level_idx * stride
                logical_end = min(logical_start + stride - 1, height - 1)
                row_start = grid_top + (height - 1 - logical_end)
                row_end = grid_top + (height - 1 - logical_start)
                _merge_if_needed(row_start, col, row_end, col)
                cell = ws.cell(row=row_start, column=col)
                cell.value = f"{axis['name']}: {level_label}"
                cell.font = label_font
                cell.alignment = y_label_alignment
                cell.border = grid_border

    for axis_idx, axis in enumerate(x_axes):
        row = x_label_start + axis_idx
        name_end_col = data_left - 1
        _merge_if_needed(row, start_col, row, name_end_col)
        name_cell = ws.cell(row=row, column=start_col)
        name_cell.value = axis["name"]
        name_cell.font = label_font
        name_cell.alignment = label_alignment
        name_cell.border = grid_border

        stride = x_strides[axis_idx]
        level_count = _axis_level_count(axis)
        period = stride * level_count
        for group_start in range(0, width, period):
            for level_idx, level_label in enumerate(axis["labels"]):
                col_start = data_left + group_start + level_idx * stride
                col_end = min(col_start + stride - 1, end_col)
                _merge_if_needed(row, col_start, row, col_end)
                cell = ws.cell(row=row, column=col_start)
                cell.value = level_label
                cell.font = x_level_font
                cell.alignment = x_level_alignment
                cell.border = grid_border

    labels_by_cell: dict[tuple[int, int], list[str]] = defaultdict(list)
    first_order_by_cell: dict[tuple[int, int], int] = {}
    for grid_row, grid_col, label, _order in placements:
        key = (grid_row, grid_col)
        labels_by_cell[key].append(label)
        first_order_by_cell[key] = min(first_order_by_cell.get(key, _order), _order)

    for (grid_row, grid_col), labels in labels_by_cell.items():
        color = _trial_sequence_color(base_color, first_order_by_cell[(grid_row, grid_col)], n_runs)
        cell = ws.cell(row=grid_top + grid_row, column=data_left + grid_col)
        cell.value = "\n".join(labels)
        cell.fill = PatternFill("solid", start_color=color, end_color=color)
        cell.font = Font(name="Arial", bold=True, color=_contrast_font_color(color), size=9)
        cell.alignment = center
        cell.border = trial_border


def _format_workbook(
    outfile: str,
    sheet_meta: dict[str, str],
    enable_spatial_visualization: bool = False,
    spatial_visualization_max_runs: int = 99,
) -> None:
    wb = load_workbook(outfile)

    row_even = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    row_odd = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"),
    )

    def _fmt(ws, hdr_bg: str, hdr_fg: str, wide_last: bool = False) -> None:
        hdr_fill = PatternFill("solid", start_color=hdr_bg, end_color=hdr_bg)
        hdr_font = Font(name="Arial", bold=True, color=hdr_fg, size=10)
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            if cell.value is not None:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = thin
        for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = row_even if ri % 2 == 0 else row_odd
            for cell in row:
                cell.font = data_font
                cell.fill = fill
                cell.alignment = center
                cell.border = thin
        for col in ws.columns:
            mw = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(mw + 3, 12)
        if wide_last and ws.max_column > 1:
            lc = get_column_letter(ws.max_column)
            ws.column_dimensions[lc].width = max(ws.column_dimensions[lc].width, 60)
        ws.freeze_panes = "B2"

    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for sname in wb.sheetnames:
        ws = wb[sname]
        tl = sheet_meta.get(sname, "")
        cat = _category(tl)
        bg, fg = CATEGORY_COLORS.get(cat, ("808080", "FFFFFF"))
        _fmt(ws, bg, fg, wide_last=(cat == "resume"))
        ws.sheet_properties.tabColor = bg

        if sname == "00_RESUME":
            notes_col = None
            for cell in ws[1]:
                if cell.value is not None and str(cell.value).strip().lower() == "notes":
                    notes_col = cell.column
                    break
            if notes_col is not None:
                for row in ws.iter_rows(min_row=2, min_col=notes_col, max_col=notes_col):
                    for cell in row:
                        cell.alignment = left_align
        elif enable_spatial_visualization:
            _add_spatial_visualization(ws, tl, spatial_visualization_max_runs, bg)

    wb.save(outfile)


# ============================================================
# FONCTION PRINCIPALE
# ============================================================

def build_doe_explorer(
    factors: Dict[str, List[Any]],
    outfile: str = "PLEX2_plans.xlsx",
    max_plan_runs: int = 30_000,
    gsd_reductions: Tuple[int, ...] = (2, 3, 4, 5),
    include_lhs: bool = True,
    lhs_samples: Optional[int] = None,
    max_full_factorial_runs: Optional[int] = None,
    enable_spatial_visualization: bool = False,
    spatial_visualization_max_runs: int = 99,
    progress_callback: Optional[Any] = None,
) -> list[dict[str, Any]]:
    if max_full_factorial_runs is not None:
        max_plan_runs = max_full_factorial_runs
    if max_plan_runs < 1:
        raise ValueError("Le seuil max de configurations par plan doit être un entier positif.")

    if spatial_visualization_max_runs < 1:
        raise ValueError("Le nombre maximal d'essais pour la visualisation doit être un entier positif.")

    def _progress(percent: int, message: str = "Génération en cours") -> None:
        if progress_callback is None:
            return
        try:
            progress_callback(max(0, min(100, int(percent))), message)
        except Exception:
            pass

    _progress(1, "Génération en cours")

    if not factors:
        raise ValueError("Aucun facteur fourni.")
    for name, lvls in factors.items():
        if not isinstance(lvls, (list, tuple)) or len(lvls) < 2:
            raise ValueError(f"Facteur '{name}' : il faut au moins 2 niveaux.")

    names = list(factors.keys())
    k = len(names)
    level_profile = [len(factors[n]) for n in names]
    all_2level = all(L == 2 for L in level_profile)

    plans: list[dict[str, Any]] = []
    omitted: list[dict[str, Any]] = []

    def skip(short_seed: str, type_label: str, reason: str, runs: int = 0) -> None:
        notes = _normalize_warning_note(type_label, str(reason), False)
        omitted.append(
            {
                "short": _build_short_label(short_seed),
                "type": type_label,
                "essais": runs,
                "colonnes": k,
                "notes": notes,
                "df": None,
                "priority": 2,
                "order": len(omitted),
            }
        )

    def _over_limit_reason(runs: int, notes: str = "") -> str:
        reason = f"{runs} essais > seuil {max_plan_runs:,}"
        clean_notes = (notes or "").strip()
        if clean_notes:
            reason = f"{reason}. {clean_notes}"
        return reason

    def add(short_seed: str, df: pd.DataFrame, type_label: str, notes: str = "") -> None:
        if isinstance(df, pd.DataFrame) and len(df) > max_plan_runs:
            skip(short_seed, type_label, _over_limit_reason(len(df), notes), len(df))
            return

        safe_df = _validate_plan_df(df, type_label)
        if len(safe_df) > max_plan_runs:
            skip(short_seed, type_label, _over_limit_reason(len(safe_df), notes), len(safe_df))
            return

        safe_notes = _normalize_warning_note(type_label, notes, True)
        plans.append(
            {
                "short": _build_short_label(short_seed),
                "type": type_label,
                "essais": len(safe_df),
                "colonnes": len(safe_df.columns) - 1,
                "notes": safe_notes,
                "df": safe_df,
                "priority": _classify_priority(type_label, safe_notes, True),
                "order": len(plans),
            }
        )

    _progress(8, "Analyse des facteurs")

    # 1. Factoriel complet
    n_full = 1
    for L in level_profile:
        n_full *= L

    if n_full <= max_plan_runs:
        add(
            "FACT_COMPLET",
            _df_full_factorial(factors),
            "Factoriel Complet",
            f"{'x'.join(str(L) for L in level_profile)} = {n_full} essais",
        )
    else:
        skip("FACT_COMPLET", "Factoriel Complet", _over_limit_reason(n_full), n_full)

    if not all_2level:
        factors_2L: Dict[str, List[Any]] = {n: [lvls[0], lvls[-1]] for n, lvls in factors.items()}
        reduced_names = [n for n in names if len(factors[n]) > 2]
        if reduced_names:
            reformat_label = (
                f"⚠ REFORMATAGE min/max : {len(reduced_names)} facteur(s) réduit(s) à 2 niveaux "
                f"({', '.join(reduced_names)}). Les niveaux intermédiaires ne sont pas utilisés dans ce plan."
            )
        else:
            reformat_label = "⚠ REFORMATAGE min/max : niveaux extrêmes utilisés."
    else:
        factors_2L = factors.copy()
        reformat_label = ""

    # 1b. Factoriel 2^k reformaté
    if not all_2level:
        n_full_2L = 2 ** k
        if n_full_2L <= max_plan_runs:
            add(
                "FACT_2L_MM",
                _df_full_factorial(factors_2L),
                "Factoriel 2^k min/max reformaté",
                f"2^{k} = {n_full_2L} essais. {reformat_label}",
            )
        else:
            skip(
                "FACT_2L_MM",
                "Factoriel 2^k min/max reformaté",
                f"2^{k} = {n_full_2L} > seuil {max_plan_runs:,}. {reformat_label}",
                n_full_2L,
            )

    # 2. Fractionnaires 2 niveaux
    if all_2level:
        try:
            from pyDOE3 import fracfact_by_res

            max_res = min(k + 1, 10)
            seen_sizes: set[int] = set()
            for res in range(max_res, 2, -1):
                try:
                    mat = fracfact_by_res(k, res)
                except Exception:
                    continue
                n = len(mat)
                if n in seen_sizes:
                    continue
                seen_sizes.add(n)
                add(
                    f"FRAC_RES{res}",
                    _df_from_pm1_2level(mat, factors),
                    f"Fractionnaire 2^({k}-p) Res.{res}",
                    f"Resolution {res}, {n} essais",
                )
        except ImportError:
            skip("FRAC_2L", "Fractionnaire 2-niveaux", "fracfact_by_res non disponible (pyDOE3 < 1.2)")

    # 2b. Fractionnaires reformatés
    if not all_2level:
        try:
            from pyDOE3 import fracfact_by_res as _fbr_mm

            seen_2L: set[int] = set()
            for res in range(min(k + 1, 10), 2, -1):
                try:
                    mat_fr = _fbr_mm(k, res)
                except Exception:
                    continue
                n_fr = len(mat_fr)
                if n_fr in seen_2L:
                    continue
                seen_2L.add(n_fr)
                add(
                    f"FRAC2L_RES{res}",
                    _df_from_pm1_2level(mat_fr, factors_2L),
                    f"Fractionnaire 2^(k-p) Res.{res} min/max reformaté",
                    f"Résolution {res}, {n_fr} essais. {reformat_label}",
                )
        except ImportError:
            skip(
                "FRAC2L",
                "Fractionnaire 2-niveaux min/max reformaté",
                "fracfact_by_res non disponible (pyDOE3 < 1.2)",
            )

    # 3. PB
    if all_2level:
        try:
            from pyDOE3 import pbdesign

            mat = pbdesign(k)
            add("PLACKETT_BURMAN", _df_from_pm1_2level(mat, factors), "Plackett-Burman", f"Criblage 2-niveaux, {len(mat)} essais")
        except Exception as e:
            skip("PLACKETT_BURMAN", "Plackett-Burman", str(e))

    # 3b. PB reformaté
    if not all_2level:
        try:
            from pyDOE3 import pbdesign as _pb_mm

            mat_pb = _pb_mm(k)
            add(
                "PB_MINMAX",
                _df_from_pm1_2level(mat_pb, factors_2L),
                "Plackett-Burman min/max reformaté",
                f"Criblage 2-niveaux, {len(mat_pb)} essais. {reformat_label}",
            )
        except Exception as e:
            skip("PB_MINMAX", "Plackett-Burman min/max reformaté", str(e))

    # 4. DSD
    quant_names = [n for n in names if _is_numeric_factor(factors[n])]
    cat_names = [n for n in names if not _is_numeric_factor(factors[n])]
    k_quant = len(quant_names)

    if k_quant >= 2:
        try:
            dsd_mat = _dsd_matrix(k_quant)
            dsd_data: Dict[str, list[Any]] = {}
            for j, n in enumerate(quant_names):
                lvls_q = factors[n]
                lo = float(min(lvls_q))
                hi = float(max(lvls_q))
                ctr = (lo + hi) / 2.0
                col_vals: list[Any] = []
                for v in dsd_mat[:, j]:
                    if abs(v) < 0.5:
                        col_vals.append(ctr)
                    elif v > 0:
                        col_vals.append(hi)
                    else:
                        col_vals.append(lo)
                dsd_data[n] = col_vals

            df_dsd = pd.DataFrame(dsd_data)
            df_dsd.insert(0, "Run", range(1, len(df_dsd) + 1))

            dsd_note = (
                f"⚠ REFORMATAGE : 3 niveaux (min, (min+max)/2, max) dérivés des valeurs min et max de chaque facteur. "
                f"2k+1 = {len(df_dsd)} essais. Construction Hadamard–Sylvester (approx. DSD). "
            )
            if cat_names:
                dsd_note += f"Facteurs non-numériques exclus : {', '.join(cat_names)} ({len(cat_names)}/{k} facteur(s))."
            if k_quant != k:
                dsd_note += f" Plan partiel sur {k_quant}/{k} facteurs quantitatifs."

            add("DSD", df_dsd, "Definitive Screening Design (DSD)", dsd_note)
        except Exception as e:
            skip("DSD", "DSD", str(e))
    else:
        skip("DSD", "DSD", f"k_quant = {k_quant} < 2 – DSD requiert ≥ 2 facteurs numériques")

    # 5. OFAT
    n_ofat_total = 1 + sum(len(factors[n]) - 1 for n in names)
    lsum = " + ".join(f"({len(factors[n])}-1)" for n in names)
    if n_ofat_total <= max_plan_runs:
        base_run = {n: factors[n][0] for n in names}
        ofat_runs: list[dict[str, Any]] = [base_run.copy()]
        for n_ofat in names:
            for lvl_ofat in factors[n_ofat][1:]:
                run_ofat = base_run.copy()
                run_ofat[n_ofat] = lvl_ofat
                ofat_runs.append(run_ofat)

        df_ofat = pd.DataFrame(ofat_runs)
        df_ofat.insert(0, "Run", range(1, len(df_ofat) + 1))
        add(
            "OFAT",
            df_ofat,
            "OFAT (One-Factor-At-a-Time)",
            f"Plan de référence. Essais = 1 + [{lsum}] = {n_ofat_total}. Base = 1er niveau de chaque facteur. Chaque facteur varie un par un, les autres restent au niveau de base.",
        )
    else:
        skip(
            "OFAT",
            "OFAT (One-Factor-At-a-Time)",
            f"Essais = 1 + [{lsum}] = {n_ofat_total} > seuil {max_plan_runs:,}",
            n_ofat_total,
        )

    # 6. Taguchi
    taguchi_list = _taguchi_all_compatible(factors)
    if taguchi_list:
        for oa_name, df_oa in taguchi_list:
            short = re.sub(r"[ ()^*]", "", oa_name)
            add(f"TAGUCHI_{short}", df_oa, f"Taguchi {oa_name}", f"{oa_name} — {len(df_oa)} essais, {k} facteurs extraits")
    else:
        skip("TAGUCHI", "Taguchi", "Aucun OA compatible pour ce profil de niveaux")

    # 7. GSD
    try:
        from pyDOE3 import gsd

        for red in gsd_reductions:
            if red <= 1:
                continue
            try:
                estimated_gsd_runs = (n_full + red - 1) // red
                if estimated_gsd_runs > max_plan_runs:
                    skip(
                        f"GSD_RED{red}",
                        f"GSD red={red}",
                        f"estimation {estimated_gsd_runs} essais > seuil {max_plan_runs:,}",
                        estimated_gsd_runs,
                    )
                    continue
                mat = gsd(level_profile, red)
                add(
                    f"GSD_RED{red}",
                    _df_from_0indexed(mat, factors),
                    f"GSD reduction={red}",
                    f"Criblage multi-niveaux, reduction={red}, {len(mat)} essais",
                )
            except Exception as e:
                skip(f"GSD_RED{red}", f"GSD red={red}", str(e))
    except ImportError:
        skip("GSD", "GSD", "gsd() non disponible dans cette version de pyDOE3")

    # 8. CCD / 9. Box-Behnken / 10. Doehlert
    numeric_factors = {n: factors[n] for n in quant_names}
    numeric_note = ""
    if cat_names:
        numeric_note = f" ⚠ REFORMATAGE : facteurs non-numériques exclus : {', '.join(cat_names)}."

    if k_quant >= 2:
        try:
            from pyDOE3 import ccdesign

            for face, label in [("ccc", "Circonscrit"), ("ccf", "Face-centre"), ("cci", "Inscrit")]:
                try:
                    n_ccd_estimated = (2 ** k_quant) + (2 * k_quant) + 2
                    if n_ccd_estimated > max_plan_runs:
                        skip(
                            f"CCD_{face.upper()}",
                            f"CCD {face}",
                            f"estimation {n_ccd_estimated} essais > seuil {max_plan_runs:,}",
                            n_ccd_estimated,
                        )
                        continue
                    mat = np.array(ccdesign(k_quant, center=(1, 1), face=face))
                    df_real, _ = _df_from_rsm(mat, numeric_factors)
                    add(
                        f"CCD_{face.upper()}",
                        df_real,
                        f"CCD {label} (valeurs reelles)",
                        f"face={face}, {len(df_real)} essais.{numeric_note}",
                    )
                except Exception as e:
                    skip(f"CCD_{face.upper()}", f"CCD {face}", str(e))
        except ImportError:
            skip("CCD", "CCD", "ccdesign() non disponible")
    else:
        skip("CCD", "CCD", "requiert au moins 2 facteurs numériques")

    if k_quant >= 3:
        try:
            from pyDOE3 import bbdesign

            n_bb_min = 2 * k_quant * (k_quant - 1)
            if n_bb_min > max_plan_runs:
                skip(
                    "BOX_BEHNKEN",
                    "Box-Behnken",
                    f"au moins {n_bb_min} essais > seuil {max_plan_runs:,}",
                    n_bb_min,
                )
            else:
                mat = np.array(bbdesign(k_quant))
                df_real, _ = _df_from_rsm(mat, numeric_factors)
                add(
                    "BOX_BEHNKEN",
                    df_real,
                    "Box-Behnken (valeurs reelles)",
                    f"{len(df_real)} essais, surface de reponse sans coin.{numeric_note}",
                )
        except Exception as e:
            skip("BOX_BEHNKEN", "Box-Behnken", str(e))
    else:
        skip("BOX_BEHNKEN", "Box-Behnken", "requiert au moins 3 facteurs numériques")

    if k_quant >= 2:
        doehlert_tasks = [
            ("doehlert_shell_design", "DOEHLERT_SHELL", "Doehlert Shell", {"num_center_points": 1}),
            ("doehlert_simplex_design", "DOEHLERT_SIMPLEX", "Doehlert Simplex", {}),
        ]
        for func_name, short_label, label, kw in doehlert_tasks:
            try:
                func = getattr(pyDOE3, func_name, None)
                if func is None:
                    exec(f"from pyDOE3 import {func_name} as func")
                mat = np.array(func(k_quant, **kw))
                df_real, _ = _df_from_rsm(mat, numeric_factors)
                add(short_label, df_real, f"{label} (valeurs reelles)", f"{len(df_real)} essais, domaine spherique.{numeric_note}")
            except Exception as e:
                skip(short_label, label, str(e))
    else:
        skip("DOEHLERT", "Doehlert", "requiert au moins 2 facteurs numériques")

    _progress(45, "Génération du Space-Filling maximin")

    # 11. Plan space-filling maximin systématique
    try:
        df_sf, note_sf = _space_filling_maximin_design(factors, max_plan_runs)
        add("SPACE_FILL_MAXIMIN", df_sf, "Space-Filling maximin", note_sf)
        if plans and plans[-1].get("short") == "SPACE_FILL_MAXIMIN":
            plans[-1]["priority"] = -1
            plans[-1]["order"] = -1
    except Exception as e:
        skip("SPACE_FILL_MAXIMIN", "Space-Filling maximin", str(e))

    # 11b. Plan space-filling orienté projections
    try:
        df_sp, note_sp = _space_filling_maxpro_design(factors, max_plan_runs)
        add("SPACE_FILL_MAXPRO", df_sp, "Space-Filling projection maximin", note_sp)
        if plans and plans[-1].get("short") == "SPACE_FILL_MAXPRO":
            plans[-1]["priority"] = -1
            plans[-1]["order"] = 0
    except Exception as e:
        skip("SPACE_FILL_MAXPRO", "Space-Filling projection maximin", str(e))

    # 12. LHS automatique et systématique
    if include_lhs:
        numeric_names = [n for n in names if _is_numeric_factor(factors[n])]
        excluded_names = [n for n in names if n not in numeric_names]
        if numeric_names:
            try:
                from pyDOE3 import lhs

                n_lhs = lhs_samples if lhs_samples else _compute_auto_lhs_samples(factors)
                if n_lhs > max_plan_runs:
                    skip("LHS_AUTO", "LHS", _over_limit_reason(n_lhs), n_lhs)
                else:
                    mat = lhs(len(numeric_names), samples=n_lhs, criterion="center")
                    data_lhs: dict[str, list[float]] = {}
                    for j, name in enumerate(numeric_names):
                        lvls = factors[name]
                        lo, hi = float(min(lvls)), float(max(lvls))
                        data_lhs[name] = [round(lo + v * (hi - lo), 6) for v in mat[:, j]]
                    df_lhs = pd.DataFrame(data_lhs)
                    df_lhs.insert(0, "Run", range(1, len(df_lhs) + 1))
                    note = f"⚠ LHS : criterion=center, {n_lhs} points (calcul automatique)."
                    if excluded_names:
                        note += f" ⚠ REFORMATAGE : facteurs non-numériques exclus : {', '.join(excluded_names)}."
                    add("LHS_AUTO", df_lhs, "Latin Hypercube Sampling", note)
            except Exception as e:
                skip("LHS_AUTO", "LHS", str(e))
        else:
            skip("LHS_AUTO", "LHS", "aucun facteur numérique disponible pour le Latin Hypercube")

    _progress(85, "Préparation du fichier Excel")

    # Tri final : nominaux -> reformatage -> omis
    generated_sorted = sorted(plans, key=lambda item: (item["priority"], item["order"], item["type"]))
    omitted_sorted = sorted(omitted, key=lambda item: (item["priority"], item["order"], item["type"]))

    used_names: set[str] = {"00_RESUME"}
    sheets: list[tuple[str, pd.DataFrame, str]] = []
    summary_rows: list[dict[str, Any]] = []
    sheet_meta: dict[str, str] = {"00_RESUME": "Resume"}

    for idx, item in enumerate(generated_sorted, start=1):
        sheet_name = _safe_name(f"{idx:02d}_{item['short']}", used_names)
        item["sheet"] = sheet_name
        sheets.append((sheet_name, item["df"], item["type"]))
        sheet_meta[sheet_name] = item["type"]
        summary_rows.append(
            {
                "Feuille": sheet_name,
                "Type": item["type"],
                "Essais": item["essais"],
                "Colonnes": item["colonnes"],
                "Notes": item["notes"],
            }
        )

    for item in omitted_sorted:
        summary_rows.append(
            {
                "Feuille": "—",
                "Type": item["type"],
                "Essais": item["essais"],
                "Colonnes": item["colonnes"],
                "Notes": item["notes"],
            }
        )

    df_resume = pd.DataFrame(summary_rows)
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        df_resume.to_excel(writer, sheet_name="00_RESUME", index=False)
        for sname, df, _ in sheets:
            df.to_excel(writer, sheet_name=sname, index=False)

    _format_workbook(
        outfile,
        sheet_meta,
        enable_spatial_visualization=enable_spatial_visualization,
        spatial_visualization_max_runs=spatial_visualization_max_runs,
    )
    _progress(100, "Terminé")
    return summary_rows


# ============================================================
# CLI (conservé pour debug interne)
# ============================================================

def _cli() -> None:
    p = argparse.ArgumentParser(description="PLEX² – plans exhaustifs dans un fichier Excel")
    p.add_argument("--k", type=int, required=True, help="Nombre de facteurs")
    p.add_argument("--l", type=int, required=True, help="Niveaux (uniforme)")
    p.add_argument("--out", default="PLEX2_plans.xlsx", help="Fichier de sortie")
    p.add_argument("--max-plan-runs", "--max-full", dest="max_plan_runs", type=int, default=30_000, help="Seuil max de configurations par plan")
    args = p.parse_args()

    import string

    names = (list(string.ascii_uppercase) + [f"X{i}" for i in range(1, 100)])[: args.k]
    factors = {n: list(range(1, args.l + 1)) for n in names}
    build_doe_explorer(factors, args.out, max_plan_runs=args.max_plan_runs, include_lhs=True)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        _cli()
    else:
        factors_level = {
            "Matériau": ["Acier", "Alu"],
            "Température": [20.0, 120.0],
            "Pression": [1.0, 2.0, 3.0],
            "Lubrification": ["Oui", "Non"],
            "Epaisseur": [0.4, 0.6],
            "Dureté": [40.0, 60.0, 70.0],
            "Ra": [0.12, 0.8, 1.6],
        }
        build_doe_explorer(factors_level, outfile="PLEX2_demo.xlsx", max_plan_runs=30_000, include_lhs=True)
